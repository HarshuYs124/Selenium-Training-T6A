import openpyxl
wb = openpyxl.Workbook()

sheetName = "sheet1"

# Create or access sheet
if sheetName in wb.sheetnames:
    ws = wb[sheetName]
else:
    ws = wb.create_sheet(sheetName)

# ws['A1'] = "Name"
# ws['B1'] = "Age"

# Save file
# wb.save("testdata.xlsx")

ws['A1']='user'
ws['B1']='passwd'

ws.append(['standard_user','secret_sauce'])
ws.append(['problem_user','secret_sauce'])
# add data to sheet
# ws.append(["Harshita", "21"])
# ws.append(["Monu", "22"])
# ws.append(["Sita", "23"])
# ws.append(["Sneha", "21"])

wb.save('sample.xlsx')

for row in ws.iter_rows(values_only=True): #to iterate in sheet
    print(row)
